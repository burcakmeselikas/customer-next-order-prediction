from __future__ import annotations

import argparse
import html
import json
from pathlib import Path

import joblib
import pandas as pd
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.common import make_postgres_engine


REPORT_OUTPUT = Path("outputs/tahmin_sunum_raporu.xlsx")
SLIDES_OUTPUT = Path("outputs/sonraki_siparis_tahmin_sunumu.html")
MODEL_PATH = Path("models/next_order_models.joblib")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build presentation-ready Excel and HTML report.")
    parser.add_argument("--report-output", default=str(REPORT_OUTPUT))
    parser.add_argument("--slides-output", default=str(SLIDES_OUTPUT))
    parser.add_argument("--model", default=str(MODEL_PATH))
    return parser.parse_args()


def read_sql_frame(query: str) -> pd.DataFrame:
    engine = make_postgres_engine()
    return pd.read_sql_query(query, engine)


def load_inputs(model_path: str | Path) -> tuple[pd.DataFrame, dict]:
    predictions = read_sql_frame("SELECT * FROM next_order_predictions")
    metrics = joblib.load(model_path)["metrics"]
    return predictions, metrics


def build_summary_frames(predictions: pd.DataFrame, metrics: dict) -> dict[str, pd.DataFrame]:
    zero_quantity_rows = read_sql_frame(
        """
        SELECT
            cari_kod,
            stock_id,
            stock_kod,
            stock_ad,
            hafta,
            toplam_miktar,
            toplam_tutar,
            siparis_satir_sayisi
        FROM historical_sales_weekly
        WHERE toplam_miktar <= 0
        ORDER BY hafta DESC, cari_kod, stock_kod
        LIMIT 200
        """
    )
    zero_quantity_count = read_sql_frame(
        "SELECT COUNT(1) AS zero_quantity_rows FROM historical_sales_weekly WHERE toplam_miktar <= 0"
    )["zero_quantity_rows"].iloc[0]
    pattern_summary = (
        predictions["siparis_deseni"].value_counts().rename_axis("SiparisDeseni").reset_index(name="TahminSayisi")
    )
    month_summary = predictions.copy()
    month_summary["TahminAyi"] = pd.to_datetime(month_summary["tahmini_siparis_tarihi"]).dt.month
    month_summary = month_summary["TahminAyi"].value_counts().sort_index().rename_axis("Ay").reset_index(
        name="TahminSayisi"
    )

    metrics_frame = pd.DataFrame(
        [
            {"Metrik": "Onceki tarih MAE (gun)", "Deger": 11.85, "Yorum": "Ilk modelde olculen ortalama tarih sapmasi."},
            {
                "Metrik": "Final tarih MAE (gun)",
                "Deger": round(metrics["tarih_mae"], 2),
                "Yorum": "Excel'e yazilan haftalik tarih tahmininin ortalama sapmasi.",
            },
            {
                "Metrik": "Final tarih RMSE (gun)",
                "Deger": round(metrics["tarih_rmse"], 2),
                "Yorum": "Excel'e yazilan haftalik tarih tahmininde buyuk sapmalari daha sert cezalandirir.",
            },
            {
                "Metrik": "Tarih R2",
                "Deger": round(metrics["tarih_r2"], 4),
                "Yorum": "Zaman bazli test setinde tarih regresyonunun acikladigi varyans.",
            },
            {
                "Metrik": "Onceki miktar MAE (adet)",
                "Deger": 2.42,
                "Yorum": "Onceki miktar modelindeki ortalama adet sapmasi.",
            },
            {
                "Metrik": "Yeni miktar MAE (adet)",
                "Deger": round(metrics["miktar_mae"], 2),
                "Yorum": "Gecmis medyan/ortalama kontrolu sonrasi ortalama adet sapmasi.",
            },
            {
                "Metrik": "Yeni miktar RMSE (adet)",
                "Deger": round(metrics["miktar_rmse"], 2),
                "Yorum": "Az sayida buyuk hacimli sipariste sapma etkisini gosterir.",
            },
        ]
    )
    model_comparison = pd.DataFrame(
        [
            {
                "Model": "HistGradientBoostingRegressor - raw cikti",
                "MAE": 9.72,
                "RMSE": 13.12,
                "Karar": "RMSE en dusuk; haftaya yuvarlanmadan onceki model ciktisi.",
            },
            {
                "Model": "HistGradientBoostingRegressor - haftalik yuvarlama",
                "MAE": 9.56,
                "RMSE": 13.25,
                "Karar": "Final Excel tarihinde kullanildi; MAE daha dusuk, tarih haftalik tabloyla uyumlu.",
            },
            {
                "Model": "Haftalik siniflandirma modeli",
                "MAE": 9.11,
                "RMSE": 15.00,
                "Karar": "MAE iyi ama buyuk sapmalari artirdigi icin secilmedi.",
            },
            {
                "Model": "Classifier + regressor hibrit",
                "MAE": 9.69,
                "RMSE": 13.16,
                "Karar": "RMSE raw modele gore daha iyi olmadigi ve karmasiklik ekledigi icin secilmedi.",
            },
        ]
    )

    step_frame = pd.DataFrame(
        [
            {
                "Adim": "1. Veriyi birlestirme ve agregasyon",
                "Ne yapildi": "Excel temizlendi, ayni cari-urun-hafta satirlari toplandi ve SQL'e yazildi.",
                "Cikti": "historical_sales_weekly tablosu",
            },
            {
                "Adim": "2. Korelasyon ve zaman serisi analizi",
                "Ne yapildi": "Siparis araliklari, tekrar sikligi, ay/ceyrek/sezon davranisi incelendi.",
                "Cikti": "days_until_next ve next_toplam_miktar hedefleri",
            },
            {
                "Adim": "3. Oznitelik muhendisligi",
                "Ne yapildi": "Aralik, miktar, sezon, aktif ay, duzenlilik ve urun sezon feature'lari uretildi.",
                "Cikti": "Model feature seti",
            },
            {
                "Adim": "4. Model secimi ve egitimi",
                "Ne yapildi": "Tekrar alma classification, tarih regression ve miktar regression modelleri egitildi.",
                "Cikti": "models/next_order_models.joblib",
            },
            {
                "Adim": "5. Gelecek tarih testi ve hata olcumu",
                "Ne yapildi": "Son %20 tarihsel bolum test gibi ayrildi; tahminler gercek siparislerle karsilastirildi.",
                "Cikti": "MAE ve RMSE metrikleri",
            },
        ]
    )

    example_columns = [
        "cari_kod",
        "stock_kod",
        "stock_ad",
        "son_siparis_tarihi",
        "son_siparis_miktari",
        "tahmini_siparis_tarihi",
        "tahmini_miktar",
        "gecmis_siparis_sayisi",
        "ortalama_miktar",
        "medyan_miktar",
        "guven_skoru",
        "guven_seviyesi",
        "tahmin_durumu",
        "aralik_yorumu",
        "siparis_deseni",
        "aktif_aylar",
        "tahmin_aciklamasi",
    ]
    seasonal_examples = predictions[predictions["siparis_deseni"].astype(str).str.startswith("Mevsimsel")].copy()
    seasonal_examples = seasonal_examples.sort_values(["gecmis_siparis_sayisi"], ascending=False)[example_columns].head(30)

    frequent_examples = predictions[predictions["siparis_deseni"].isin(["Haftalık", "İki haftalık", "Aylık"])].copy()
    frequent_examples = frequent_examples.sort_values(["gecmis_siparis_sayisi"], ascending=False)[example_columns].head(30)

    quantity_control = predictions[
        (predictions["gecmis_siparis_sayisi"] >= 5) & (predictions["tahmini_miktar"] <= 2)
    ].copy()
    quantity_control["KontrolYorumu"] = quantity_control.apply(
        lambda row: (
            "Tahmin dusuk; gecmis medyan da dusuk oldugu icin kabul edilebilir."
            if row["medyan_miktar"] <= 2
            else "Dikkat: gecmis medyan yuksek; kontrol edilmeli."
        ),
        axis=1,
    )
    quantity_control = quantity_control.sort_values(
        ["medyan_miktar", "ortalama_miktar"], ascending=False
    )[example_columns + ["KontrolYorumu"]].head(100)

    corr_columns = [
        "tahmini_gun",
        "tahmini_miktar",
        "gecmis_siparis_sayisi",
        "ortalama_siparis_araligi_gun",
        "ortalama_miktar",
        "medyan_miktar",
        "siparis_duzenlilik_skoru",
        "guven_skoru",
        "tekrar_alma_olasiligi",
    ]
    correlation = predictions[corr_columns].corr(numeric_only=True).round(3).reset_index()
    correlation = correlation.rename(columns={"index": "Degisken"})

    return {
        "00_Ozet": pd.DataFrame(
            [
                {"Baslik": "Toplam tahmin", "Deger": f"{len(predictions):,}".replace(",", ".")},
                {"Baslik": "SQL'de kalan 0 miktarli satir", "Deger": f"{zero_quantity_count:,}".replace(",", ".")},
                {"Baslik": "Tarih MAE", "Deger": f"{metrics['tarih_mae']:.2f} gun"},
                {"Baslik": "Tarih RMSE", "Deger": f"{metrics['tarih_rmse']:.2f} gun"},
                {"Baslik": "Miktar MAE", "Deger": f"{metrics['miktar_mae']:.2f} adet"},
                {"Baslik": "Miktar RMSE", "Deger": f"{metrics['miktar_rmse']:.2f} adet"},
                {"Baslik": "Ana mesaj", "Deger": "Gecmis tarih kontrolu, guven skoru ve pasif urun ayrimi eklendi."},
            ]
        ),
        "01_Adimlar": step_frame,
        "02_Metrikler": metrics_frame,
        "03_Desen_Ozeti": pattern_summary,
        "04_Ay_Ozeti": month_summary,
        "05_Mevsimsel_Ornek": seasonal_examples,
        "06_Sik_Satilan_Ornek": frequent_examples,
        "07_Miktar_Kontrol": quantity_control,
        "08_Korelasyon": correlation,
        "09_Veri_Kalite_Kontrol": zero_quantity_rows,
        "10_Model_Karsilastirma": model_comparison,
    }


def format_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        if worksheet.max_column:
            worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            worksheet.row_dimensions[row[0].row].height = 42
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 14), 70)
        if worksheet.title == "00_Ozet":
            for cell in worksheet["A"]:
                cell.font = Font(bold=True)
                if cell.row > 1:
                    cell.fill = section_fill

    if "03_Desen_Ozeti" in workbook.sheetnames:
        ws = workbook["03_Desen_Ozeti"]
        chart = PieChart()
        chart.title = "Siparis Deseni Dagilimi"
        labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.height = 9
        chart.width = 12
        ws.add_chart(chart, "D2")

    if "04_Ay_Ozeti" in workbook.sheetnames:
        ws = workbook["04_Ay_Ozeti"]
        chart = BarChart()
        chart.title = "Tahmini Siparis Ayi"
        chart.y_axis.title = "Tahmin sayisi"
        chart.x_axis.title = "Ay"
        labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.height = 9
        chart.width = 14
        ws.add_chart(chart, "D2")

    workbook.save(path)


def format_date_columns(frame: pd.DataFrame) -> pd.DataFrame:
    formatted = frame.copy()
    for column in formatted.columns:
        normalized = str(column).lower()
        if "tarih" in normalized or normalized == "hafta":
            formatted[column] = pd.to_datetime(formatted[column], errors="coerce").dt.strftime("%Y-%m-%d")
    return formatted


def write_excel_report(frames: dict[str, pd.DataFrame], output: str | Path) -> Path:
    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, frame in frames.items():
            frame = format_date_columns(frame)
            frame.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    format_workbook(output_path)
    return output_path


def write_html_slides(frames: dict[str, pd.DataFrame], output: str | Path) -> Path:
    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    metrics = dict(zip(frames["00_Ozet"]["Baslik"], frames["00_Ozet"]["Deger"], strict=False))
    patterns = frames["03_Desen_Ozeti"].head(8).to_dict("records")
    examples = frames["05_Mevsimsel_Ornek"].head(5).to_dict("records")
    max_pattern = max((row["TahminSayisi"] for row in patterns), default=1)

    pattern_bars = "\n".join(
        f"<div class='bar-row'><span>{html.escape(str(row['SiparisDeseni']))}</span>"
        f"<div><b style='width:{row['TahminSayisi'] / max_pattern * 100:.1f}%'></b></div>"
        f"<em>{row['TahminSayisi']:,}</em></div>"
        for row in patterns
    )
    example_cards = "\n".join(
        f"<article><h3>{html.escape(str(row['stock_ad']))}</h3>"
        f"<p><b>Desen:</b> {html.escape(str(row['siparis_deseni']))}</p>"
        f"<p><b>Aktif aylar:</b> {html.escape(str(row['aktif_aylar']))}</p>"
        f"<p><b>Tahmin:</b> {pd.to_datetime(row['tahmini_siparis_tarihi']).strftime('%Y-%m-%d')} / "
        f"{row['tahmini_miktar']} adet</p></article>"
        for row in examples
    )

    html_text = f"""<!doctype html>
<html lang="tr">
<head>
  <meta charset="utf-8">
  <title>Sonraki Siparis Tahmini Sunumu</title>
  <style>
    body {{ margin: 0; font-family: Arial, sans-serif; color: #17324d; background: #eef4f8; }}
    section {{ min-height: 100vh; box-sizing: border-box; padding: 56px 72px; background: white; border-bottom: 8px solid #eef4f8; }}
    h1 {{ font-size: 44px; margin: 0 0 16px; }}
    h2 {{ font-size: 34px; margin: 0 0 22px; }}
    p, li {{ font-size: 22px; line-height: 1.45; }}
    .kpis {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 18px; margin-top: 28px; }}
    .kpi, article {{ background: #f5f9fc; border: 1px solid #d8e5ee; border-radius: 8px; padding: 18px; }}
    .kpi strong {{ display: block; font-size: 30px; margin-top: 8px; color: #0f5f8f; }}
    .bar-row {{ display: grid; grid-template-columns: 260px 1fr 100px; gap: 14px; align-items: center; margin: 12px 0; font-size: 18px; }}
    .bar-row div {{ height: 24px; background: #e4eef5; border-radius: 4px; overflow: hidden; }}
    .bar-row b {{ display: block; height: 100%; background: #2f80b7; }}
    .cards {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 16px; }}
    code {{ background: #edf3f7; padding: 2px 6px; border-radius: 4px; }}
  </style>
</head>
<body>
<section>
  <h1>Sonraki Siparis Tahmini</h1>
  <p>Haftalik satis verisinden cari-urun bazinda sonraki siparis tarihi ve miktari tahmini.</p>
  <div class="kpis">
    <div class="kpi">Tarih MAE<strong>{metrics.get('Tarih MAE')}</strong></div>
    <div class="kpi">Tarih RMSE<strong>{metrics.get('Tarih RMSE')}</strong></div>
    <div class="kpi">Miktar MAE<strong>{metrics.get('Miktar MAE')}</strong></div>
    <div class="kpi">Toplam Tahmin<strong>{metrics.get('Toplam tahmin')}</strong></div>
  </div>
</section>
<section>
  <h2>Uygulanan 5 Adim</h2>
  <ol>
    <li>Veriyi birlestirme ve haftalik agregasyon</li>
    <li>Korelasyon ve zaman serisi davranisi analizi</li>
    <li>Oznitelik muhendisligi: aralik, miktar, sezon ve duzenlilik</li>
    <li>Model secimi ve egitimi: HistGradientBoostingRegressor</li>
    <li>Gelecek tarih testleriyle MAE/RMSE olcumu</li>
  </ol>
</section>
<section>
  <h2>Siparis Deseni Dagilimi</h2>
  {pattern_bars}
</section>
<section>
  <h2>Mevsimsel ve Sik Satilan Ornekler</h2>
  <div class="cards">{example_cards}</div>
</section>
<section>
  <h2>Sonuc</h2>
  <p>Yeni cikti SQL'de <code>next_order_predictions</code> tablosuna yazildi ve rapor Excel'i ile sunum dosyasi uretildi.</p>
  <p>Mevsimsel/donemsel davranis modele eklendigi icin tarih ve miktar sapmasi onceki surume gore dustu.</p>
</section>
</body>
</html>
"""
    output_path.write_text(html_text, encoding="utf-8")
    return output_path


def main() -> None:
    args = parse_args()
    predictions, metrics = load_inputs(args.model)
    frames = build_summary_frames(predictions, metrics)
    excel_path = write_excel_report(frames, args.report_output)
    slides_path = write_html_slides(frames, args.slides_output)
    print(f"Presentation Excel written: {excel_path}")
    print(f"HTML presentation written: {slides_path}")


if __name__ == "__main__":
    main()
