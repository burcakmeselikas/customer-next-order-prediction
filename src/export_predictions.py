from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.common import make_postgres_engine, table_reference


STATUS_VALID = "Ge\u00e7erli Tahmin"
STATUS_OVERDUE = "Gecikmi\u015f / Overdue"
STATUS_PASSIVE = "Pasif \u00dcr\u00fcn"
STATUS_INSUFFICIENT = "Yetersiz Ge\u00e7mi\u015f"
STATUS_LOW_PROBABILITY = "D\u00fc\u015f\u00fck Olas\u0131l\u0131kl\u0131 Tahmin"

STATUS_PRIORITY = {
    STATUS_VALID: 0,
    STATUS_OVERDUE: 1,
    STATUS_LOW_PROBABILITY: 2,
    STATUS_INSUFFICIENT: 3,
    STATUS_PASSIVE: 4,
}

EXCEL_COLUMN_MAP = {
    "cari_id": "CariId",
    "cari_kod": "CariKod",
    "stock_id": "StockId",
    "stock_kod": "StockKod",
    "stock_ad": "StockAd",
    "son_siparis_tarihi": "SonSiparisTarihi",
    "son_siparis_miktari": "SonSiparisMiktari",
    "gecmis_siparis_sayisi": "GecmisSiparisSayisi",
    "son_siparisten_gecen_gun": "SonSiparistenGecenGun",
    "tahmini_siparis_tarihi": "TahminiSiparisTarihi",
    "tahmini_gun": "TahminiGun",
    "tahmini_miktar": "TahminiMiktar",
    "tekrar_alma_olasiligi": "TekrarAlmaOlasiligi",
    "guven_skoru": "GuvenSkoru",
    "guven_seviyesi": "GuvenSeviyesi",
    "tahmin_durumu": "TahminDurumu",
    "tahmin_aciklamasi": "TahminAciklamasi",
    "ortalama_siparis_araligi_gun": "OrtalamaSiparisAraligiGun",
    "medyan_siparis_araligi_gun": "MedyanSiparisAraligiGun",
    "ortalama_miktar": "OrtalamaMiktar",
    "medyan_miktar": "MedyanMiktar",
    "son3_siparis_ortalamasi": "Son3SiparisOrtalamasi",
    "son6_siparis_ortalamasi": "Son6SiparisOrtalamasi",
    "siparis_araligi_std": "SiparisAraligiStd",
    "miktar_std": "MiktarStd",
    "miktar_p95": "MiktarP95",
    "aktif_aylar": "AktifAylar",
    "siparis_duzenlilik_skoru": "SiparisDuzenlilikSkoru",
    "aralik_yorumu": "AralikYorumu",
    "siparis_deseni": "SiparisDeseni",
    "tarih_tahmin_yontemi": "TarihTahminYontemi",
    "model_tahmini_siparis_tarihi": "ModelTahminiSiparisTarihi",
    "baseline_tahmini_siparis_tarihi": "BaselineTahminiSiparisTarihi",
    "tarih_yayma_uygulandi": "TarihYaymaUygulandi",
    "ham_tahmini_siparis_tarihi": "HamTahminiSiparisTarihi",
    "ham_tahmin_gecmise_dustu": "HamTahminGecmiseDustu",
    "aksiyon_tahmini": "AksiyonTahmini",
}

MAIN_OUTPUT_COLUMNS = [
    "CariId",
    "CariKod",
    "StockId",
    "StockKod",
    "StockAd",
    "SonSiparisTarihi",
    "SonSiparisMiktari",
    "GecmisSiparisSayisi",
    "SonSiparistenGecenGun",
    "TahminiSiparisTarihi",
    "TahminiGun",
    "TahminiMiktar",
    "TekrarAlmaOlasiligi",
    "GuvenSkoru",
    "GuvenSeviyesi",
    "TahminDurumu",
    "TahminAciklamasi",
    "OrtalamaSiparisAraligiGun",
    "MedyanSiparisAraligiGun",
    "OrtalamaMiktar",
    "MedyanMiktar",
    "Son3SiparisOrtalamasi",
    "Son6SiparisOrtalamasi",
    "SiparisAraligiStd",
    "MiktarStd",
    "MiktarP95",
    "AktifAylar",
    "SiparisDuzenlilikSkoru",
    "AralikYorumu",
    "SiparisDeseni",
    "TarihTahminYontemi",
    "ModelTahminiSiparisTarihi",
    "BaselineTahminiSiparisTarihi",
    "TarihYaymaUygulandi",
    "HamTahminiSiparisTarihi",
    "HamTahminGecmiseDustu",
    "AksiyonTahmini",
]

ACTION_OUTPUT_COLUMNS = ["OncelikSkoru"] + MAIN_OUTPUT_COLUMNS


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export PostgreSQL prediction table to Excel.")
    parser.add_argument("--table", default="next_order_predictions")
    parser.add_argument("--schema", default=None)
    parser.add_argument("--output", default="outputs/next_order_predictions_from_postgres.xlsx")
    return parser.parse_args()


def sort_predictions(predictions: pd.DataFrame) -> pd.DataFrame:
    frame = predictions.copy()
    if "tahmin_durumu" in frame.columns:
        frame["_tahmin_durumu_sira"] = frame["tahmin_durumu"].map(STATUS_PRIORITY).fillna(99)
        sort_columns = ["_tahmin_durumu_sira"]
        ascending = [True]
        if "tahmini_siparis_tarihi" in frame.columns:
            sort_columns.append("tahmini_siparis_tarihi")
            ascending.append(True)
        if "guven_skoru" in frame.columns:
            sort_columns.append("guven_skoru")
            ascending.append(False)
        for column in ["cari_kod", "stock_kod", "stock_ad"]:
            if column in frame.columns:
                sort_columns.append(column)
                ascending.append(True)
        frame = frame.sort_values(sort_columns, ascending=ascending).drop(columns="_tahmin_durumu_sira")
    elif {"cari_kod", "stock_kod", "stock_ad"}.issubset(frame.columns):
        frame = frame.sort_values(["cari_kod", "stock_kod", "stock_ad"])
    return frame.reset_index(drop=True)


def sort_action_predictions(predictions: pd.DataFrame) -> pd.DataFrame:
    frame = predictions.copy()
    sort_columns: list[str] = []
    ascending: list[bool] = []
    if "tahmini_siparis_tarihi" in frame.columns:
        sort_columns.append("tahmini_siparis_tarihi")
        ascending.append(True)
    if "guven_skoru" in frame.columns:
        sort_columns.append("guven_skoru")
        ascending.append(False)
    if "tekrar_alma_olasiligi" in frame.columns:
        sort_columns.append("tekrar_alma_olasiligi")
        ascending.append(False)
    for column in ["cari_kod", "stock_kod", "stock_ad"]:
        if column in frame.columns:
            sort_columns.append(column)
            ascending.append(True)
    if sort_columns:
        frame = frame.sort_values(sort_columns, ascending=ascending)
    return frame.reset_index(drop=True)


def read_predictions(table: str, schema: str | None) -> pd.DataFrame:
    engine = make_postgres_engine()
    query = f"SELECT * FROM {table_reference(table, schema)}"
    return sort_predictions(pd.read_sql_query(query, engine))


def _format_date_columns(frame: pd.DataFrame) -> pd.DataFrame:
    formatted = frame.copy()
    for column in formatted.columns:
        column_text = str(column)
        normalized = column_text.lower()
        is_date_column = (
            column_text.endswith("Tarihi")
            or column_text.endswith("_tarihi")
            or normalized.endswith("_date")
            or normalized == "hafta"
        )
        if is_date_column:
            dates = pd.to_datetime(formatted[column], errors="coerce")
            formatted[column] = dates.dt.strftime("%Y-%m-%d").where(dates.notna(), "")
    return formatted


def _rename_predictions_for_excel(predictions: pd.DataFrame) -> pd.DataFrame:
    excel_predictions = predictions.rename(columns=EXCEL_COLUMN_MAP)
    excel_predictions = _format_date_columns(excel_predictions)
    output_columns = ACTION_OUTPUT_COLUMNS if "OncelikSkoru" in excel_predictions.columns else MAIN_OUTPUT_COLUMNS
    ordered = [column for column in output_columns if column in excel_predictions.columns]
    remaining = [column for column in excel_predictions.columns if column not in ordered]
    return excel_predictions[ordered + remaining]


def _prepare_action_sheet(predictions: pd.DataFrame) -> pd.DataFrame:
    action = sort_action_predictions(predictions)
    if action.empty:
        return _empty_prediction_sheet(include_priority=True)

    action["oncelik_skoru"] = (
        action["tekrar_alma_olasiligi"].astype(float) * 0.45
        + (action["guven_skoru"].astype(float) / 100) * 0.35
        + action["siparis_duzenlilik_skoru"].fillna(0).astype(float) * 0.20
    ).round(4)

    high_quantity_mask = (
        action["tahmini_miktar"].notna()
        & action["miktar_p95"].notna()
        & (action["tahmini_miktar"].astype(float) >= action["miktar_p95"].astype(float) * 0.90)
    )
    if high_quantity_mask.any():
        warning = "Y\u00fcksek Miktar Kontrol Edilmeli"
        action.loc[high_quantity_mask, "tahmin_aciklamasi"] = (
            action.loc[high_quantity_mask, "tahmin_aciklamasi"].fillna("").astype(str).str.rstrip()
            + " "
            + warning
            + "."
        )

    action = action.rename(columns={"oncelik_skoru": "OncelikSkoru"})
    return _rename_predictions_for_excel(action)


def _metric_label(key: str) -> str:
    labels = {
        "repeat_horizon_days": "Tekrar alma siniflandirma ufku (gun)",
        "data_max_hafta": "Veri seti maksimum Hafta",
        "classifier_training_rows": "Classification egitim satiri",
        "classifier_validation_rows": "Classification test satiri",
        "classifier_split_date": "Classification zaman ayrimi",
        "classifier_accuracy": "Classification accuracy",
        "classifier_precision": "Precision",
        "classifier_recall": "Recall",
        "classifier_f1": "F1 Score",
        "classifier_roc_auc": "Classification ROC AUC",
        "classifier_average_precision": "Classification average precision",
        "classifier_positive_class_ratio": "Positive class orani",
        "confusion_matrix_tn": "Confusion Matrix TN",
        "confusion_matrix_fp": "Confusion Matrix FP",
        "confusion_matrix_fn": "Confusion Matrix FN",
        "confusion_matrix_tp": "Confusion Matrix TP",
        "training_rows": "Regression egitim satiri",
        "validation_rows": "Regression test satiri",
        "regression_split_date": "Regression zaman ayrimi",
        "date_prediction_method": "Kullanilan tarih tahmin yontemi",
        "baseline_tarih_mae": "Baseline tarih MAE",
        "raw_model_tarih_mae": "Raw model tarih MAE",
        "baseline_miktar_mae": "Baseline miktar MAE",
        "raw_tarih_model_iyilesme_yuzde": "Raw tarih modelinin baseline'a gore iyilesmesi (%)",
        "kullanilan_tarih_iyilesme_yuzde": "Kullanilan tarih yontemi baseline'a gore iyilesme (%)",
        "miktar_model_iyilesme_yuzde": "Miktar modelinin baseline'a gore iyilesmesi (%)",
        "tarih_mae": "Tarih MAE (gun)",
        "tarih_rmse": "Tarih RMSE (gun)",
        "miktar_mae": "Miktar MAE",
        "miktar_rmse": "Miktar RMSE",
        "tarih_r2": "Tarih R2",
        "miktar_r2": "Miktar R2",
        "tarih_smape": "Tarih sMAPE (%)",
        "miktar_smape": "Miktar sMAPE (%)",
        "prediction_rows": "Toplam cari-urun satiri",
        "actionable_prediction_rows": "Aksiyon tahmini satiri",
        "overdue_rows": "Gecikmis / overdue satiri",
        "passive_product_rows": "Pasif urun satiri",
        "low_probability_rows": "Dusuk olasilikli satir",
    }
    return labels.get(key, key)


def _metrics_to_frame(metrics: dict[str, Any] | None) -> pd.DataFrame:
    if not metrics:
        return pd.DataFrame(columns=["Metrik", "Deger"])
    rows = [{"Metrik": _metric_label(key), "Deger": value} for key, value in metrics.items()]
    if metrics.get("date_prediction_method") == "model_regression":
        rows.append(
            {
                "Metrik": "Yorum",
                "Deger": "Tarih modeli baseline'dan daha iyi oldugu icin kullanildi.",
            }
        )
    else:
        rows.append(
            {
                "Metrik": "Yorum",
                "Deger": "Tarih modeli baseline'dan iyi olmadigi icin medyan aralik baseline'i kullanildi.",
            }
        )
    rows.extend(
        [
            {
                "Metrik": "Yorum",
                "Deger": "Miktar modelinde RMSE, buyuk siparislerden etkilenmektedir.",
            },
            {
                "Metrik": "Yorum",
                "Deger": "Aksiyon tahminleri yalnizca guvenilir ve aktif cari-urun ciftlerinden olusmaktadir.",
            },
        ]
    )
    return pd.DataFrame(rows)


def _build_guven_ozeti(predictions: pd.DataFrame) -> pd.DataFrame:
    required = {"guven_seviyesi", "tahmin_durumu"}
    if not required.issubset(predictions.columns):
        return pd.DataFrame(columns=["GuvenSeviyesi", "TahminDurumu", "SatirSayisi"])

    aggregations: dict[str, tuple[str, str]] = {"SatirSayisi": ("tahmin_durumu", "size")}
    if "aksiyon_tahmini" in predictions.columns:
        aggregations["AksiyonSatiri"] = ("aksiyon_tahmini", "sum")
    if "guven_skoru" in predictions.columns:
        aggregations["OrtalamaGuvenSkoru"] = ("guven_skoru", "mean")
    if "tekrar_alma_olasiligi" in predictions.columns:
        aggregations["OrtalamaTekrarAlmaOlasiligi"] = ("tekrar_alma_olasiligi", "mean")

    summary = (
        predictions.groupby(["guven_seviyesi", "tahmin_durumu"], dropna=False)
        .agg(**aggregations)
        .reset_index()
        .rename(columns={"guven_seviyesi": "GuvenSeviyesi", "tahmin_durumu": "TahminDurumu"})
    )
    for column in ["OrtalamaGuvenSkoru", "OrtalamaTekrarAlmaOlasiligi"]:
        if column in summary.columns:
            summary[column] = summary[column].round(4)
    summary["_sira"] = summary["TahminDurumu"].map(STATUS_PRIORITY).fillna(99)
    return summary.sort_values(["_sira", "GuvenSeviyesi"]).drop(columns="_sira").reset_index(drop=True)


def _empty_prediction_sheet(include_priority: bool = False) -> pd.DataFrame:
    return pd.DataFrame(columns=ACTION_OUTPUT_COLUMNS if include_priority else MAIN_OUTPUT_COLUMNS)


def _action_filter(predictions: pd.DataFrame, metrics: dict[str, Any] | None = None) -> pd.Series:
    if predictions.empty:
        return pd.Series(False, index=predictions.index)
    if "aksiyon_tahmini" in predictions.columns:
        return predictions["aksiyon_tahmini"].fillna(False).astype(bool)

    max_date = pd.to_datetime((metrics or {}).get("data_max_hafta"), errors="coerce")
    future_ok = pd.Series(True, index=predictions.index)
    if not pd.isna(max_date) and "tahmini_siparis_tarihi" in predictions.columns:
        future_ok = pd.to_datetime(predictions["tahmini_siparis_tarihi"], errors="coerce") > max_date
    return (
        predictions["tahmin_durumu"].eq(STATUS_VALID)
        & predictions["tekrar_alma_olasiligi"].ge(0.50)
        & predictions["guven_skoru"].ge(60)
        & predictions["gecmis_siparis_sayisi"].ge(3)
        & predictions["son_siparisten_gecen_gun"].le(90)
        & future_ok
    )


def _metric_value(metrics: dict[str, Any] | None, key: str, default: Any = None) -> Any:
    if not metrics:
        return default
    return metrics.get(key, default)


def _build_manager_summary(
    predictions: pd.DataFrame,
    action_predictions: pd.DataFrame,
    metrics: dict[str, Any] | None,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = [
        {"Baslik": "Toplam cari-urun satiri", "Deger": int(len(predictions))},
        {"Baslik": "Aksiyon tahmini sayisi", "Deger": int(len(action_predictions))},
        {"Baslik": "Overdue sayisi", "Deger": int((predictions["tahmin_durumu"] == STATUS_OVERDUE).sum())},
        {"Baslik": "Pasif urun sayisi", "Deger": int((predictions["tahmin_durumu"] == STATUS_PASSIVE).sum())},
        {
            "Baslik": "Yetersiz gecmis sayisi",
            "Deger": int((predictions["tahmin_durumu"] == STATUS_INSUFFICIENT).sum()),
        },
        {
            "Baslik": "Dusuk olasilikli satir sayisi",
            "Deger": int((predictions["tahmin_durumu"] == STATUS_LOW_PROBABILITY).sum()),
        },
        {"Baslik": "Tarih MAE", "Deger": _metric_value(metrics, "tarih_mae")},
        {"Baslik": "Tarih RMSE", "Deger": _metric_value(metrics, "tarih_rmse")},
        {"Baslik": "Miktar MAE", "Deger": _metric_value(metrics, "miktar_mae")},
        {"Baslik": "Miktar RMSE", "Deger": _metric_value(metrics, "miktar_rmse")},
        {"Baslik": "ROC AUC", "Deger": _metric_value(metrics, "classifier_roc_auc")},
    ]

    max_date = pd.to_datetime(_metric_value(metrics, "data_max_hafta"), errors="coerce")
    if pd.isna(max_date):
        max_date = pd.to_datetime(action_predictions["tahmini_siparis_tarihi"], errors="coerce").min()
        if pd.isna(max_date):
            max_date = pd.Timestamp.today().normalize()

    action_dates = pd.to_datetime(action_predictions.get("tahmini_siparis_tarihi"), errors="coerce")
    for week_number in range(1, 5):
        week_date = max_date + pd.Timedelta(days=7 * week_number)
        count = int((action_dates == week_date).sum())
        rows.append({"Baslik": f"En yakin {week_number}. hafta tahmin sayisi ({week_date.date()})", "Deger": count})

    return pd.DataFrame(rows)


def build_prediction_summaries(
    predictions: pd.DataFrame,
    metrics: dict[str, Any] | None = None,
    analytics: dict[str, pd.DataFrame] | None = None,
) -> dict[str, pd.DataFrame]:
    action_mask = _action_filter(predictions, metrics)
    action_predictions = sort_action_predictions(predictions[action_mask].copy())
    overdue = sort_predictions(predictions[predictions["tahmin_durumu"] == STATUS_OVERDUE].copy())
    passive = sort_predictions(predictions[predictions["tahmin_durumu"] == STATUS_PASSIVE].copy())
    insufficient = sort_predictions(predictions[predictions["tahmin_durumu"] == STATUS_INSUFFICIENT].copy())
    low_probability = sort_predictions(predictions[predictions["tahmin_durumu"] == STATUS_LOW_PROBABILITY].copy())

    analytics = analytics or {}
    return {
        "aksiyon_tahminleri": _prepare_action_sheet(action_predictions),
        "yonetici_ozeti": _build_manager_summary(predictions, action_predictions, metrics),
        "gecikmis_overdue_urunler": (
            _rename_predictions_for_excel(overdue) if not overdue.empty else _empty_prediction_sheet()
        ),
        "pasif_urunler": _rename_predictions_for_excel(passive) if not passive.empty else _empty_prediction_sheet(),
        "yetersiz_gecmis": (
            _rename_predictions_for_excel(insufficient) if not insufficient.empty else _empty_prediction_sheet()
        ),
        "dusuk_olasilikli_tahminler": (
            _rename_predictions_for_excel(low_probability) if not low_probability.empty else _empty_prediction_sheet()
        ),
        "model_metrics": _metrics_to_frame(metrics),
        "guven_ozeti": _build_guven_ozeti(predictions),
        "hata_analizi": analytics.get("hata_analizi", pd.DataFrame()),
        "feature_importance": analytics.get("feature_importance", pd.DataFrame()),
    }


def _autosize_columns(writer: pd.ExcelWriter, sample_rows: int = 1000) -> None:
    for worksheet in writer.book.worksheets:
        max_row = min(worksheet.max_row, sample_rows)
        for column_index in range(1, worksheet.max_column + 1):
            max_length = 0
            for row_index in range(1, max_row + 1):
                value = worksheet.cell(row=row_index, column=column_index).value
                if value is not None:
                    max_length = max(max_length, len(str(value)))
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 10), 60)

        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        worksheet.freeze_panes = "A2"
        if worksheet.max_column and worksheet.max_row:
            worksheet.auto_filter.ref = worksheet.dimensions

        for header_cell in worksheet[1]:
            if header_cell.value in {"TahminAciklamasi", "AktifAylar"}:
                worksheet.column_dimensions[get_column_letter(header_cell.column)].width = 70
            elif header_cell.value == "StockAd":
                worksheet.column_dimensions[get_column_letter(header_cell.column)].width = 45


def export_predictions_to_excel(
    predictions: pd.DataFrame,
    output: str | Path,
    metrics: dict[str, Any] | None = None,
    include_summaries: bool = True,
    analytics: dict[str, pd.DataFrame] | None = None,
) -> Path:
    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    sorted_predictions = sort_predictions(predictions)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if include_summaries:
            for sheet_name, frame in build_prediction_summaries(sorted_predictions, metrics, analytics).items():
                formatted = _format_date_columns(frame)
                formatted.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        else:
            action_predictions = sorted_predictions[_action_filter(sorted_predictions, metrics)].copy()
            _rename_predictions_for_excel(action_predictions).to_excel(
                writer,
                sheet_name="aksiyon_tahminleri",
                index=False,
            )
        _autosize_columns(writer)

    return output_path


def main() -> None:
    args = parse_args()
    predictions = read_predictions(args.table, args.schema)
    output_path = export_predictions_to_excel(predictions, args.output)
    print(f"Exported {len(predictions):,} prediction rows from {table_reference(args.table, args.schema)}.")
    print(f"Excel written: {output_path}")


if __name__ == "__main__":
    main()
